"""
Админ-панель для модуля «Сметы» (app_outlay).
"""

import os
import json
import tempfile

from decimal import Decimal
from openpyxl import load_workbook

from django.db import transaction
from django.db.models import Count
from django.core.cache import cache
from django.urls import reverse, path
from django.contrib import admin, messages
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponseRedirect, HttpResponse


from app_outlay.forms import GroupFormSet, LinkFormSet
from app_outlay.models import (
    Estimate,
    Group,
    GroupTechnicalCardLink,
    EstimateOverheadCostLink,
)
from app_technical_cards.models import TechnicalCard as _TC
from app_estimate_imports.services.schema_service import SchemaService as _SS


# ---------- Импорты для ENDPOINTS  ----------
from app_outlay.views.estimate_calc_view.utils_calc import (
    calc_for_tc,
    _base_costs_live,
    _dec,
)


# ---------- Для чтения листов  ----------
# это бы потом вынести от сюда, например в утилиты
def _xlsx_cache_key(path: str, sheet_index: int) -> str:
    try:
        mtime = int(os.path.getmtime(path))
    except Exception:
        mtime = 0
    return f"outlay:xlsx:{path}:{mtime}:sheet:{sheet_index}"


def _load_full_sheet_rows(xlsx_path: str, sheet_index: int) -> list[dict]:
    """
    Возвращает ВСЕ строки листа в виде [{'row_index': int, 'cells': [..]}, ..]
    row_index — 1-based как в разметке групп.
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb.worksheets[sheet_index]
    except IndexError:
        ws = wb.active

    rows = []
    # оценим ширину по первой непустой сотне строк
    max_cols = 0
    sample = 0
    for r in ws.iter_rows(min_row=1, max_row=min(200, ws.max_row), values_only=True):
        if any(c not in (None, "", " ") for c in r):
            max_cols = max(max_cols, len(r))
        sample += 1
        if sample >= 200:
            break
    if max_cols <= 0:
        max_cols = ws.max_column or 1

    # теперь читаем всё
    for idx, r in enumerate(
        ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1
    ):
        cells = list(r)[:max_cols]
        # нормализация текста
        norm = [(str(c).strip() if c is not None else "") for c in cells]
        rows.append({"row_index": idx, "cells": norm})
    wb.close()
    return rows


def _load_full_sheet_rows_cached(
    xlsx_path: str, sheet_index: int, ttl: int = 600
) -> list[dict]:
    """
    Обёртка над `_load_full_sheet_rows` с кешированием результатов.

    :param xlsx_path: путь к файлу .xlsx
    :param sheet_index: индекс листа
    :param ttl: время жизни кеша в секундах (дефолт: 10 минут)
    :return: список словарей строк листа
    """
    key = _xlsx_cache_key(xlsx_path, sheet_index)
    cached = cache.get(key)
    if cached is not None:
        return cached
    rows = _load_full_sheet_rows(xlsx_path, sheet_index)
    cache.set(key, rows, ttl)
    return rows


# ---------- АДМИНКИ ----------

ROLE_TITLES = {
    "NAME_OF_WORK": "НАИМЕНОВАНИЕ РАБОТ/ТК",
    "UNIT": "ЕД. ИЗМ.",
    "QTY": "КОЛ-ВО",
    "UNIT_PRICE_OF_MATERIAL": "ЦЕНА МАТ/ЕД",
    "UNIT_PRICE_OF_WORK": "ЦЕНА РАБОТЫ/ЕД",
    "UNIT_PRICE_OF_MATERIALS_AND_WORKS": "ЦЕНА МАТ+РАБ/ЕД",
    "PRICE_FOR_ALL_MATERIAL": "ИТОГО МАТЕРИАЛ",
    "PRICE_FOR_ALL_WORK": "ИТОГО РАБОТА",
    "TOTAL_PRICE": "ОБЩАЯ ЦЕНА",
}
OPTIONAL_ROLE_IDS = [
    "UNIT_PRICE_OF_MATERIAL",
    "UNIT_PRICE_OF_WORK",
    "UNIT_PRICE_OF_MATERIALS_AND_WORKS",
    "PRICE_FOR_ALL_MATERIAL",
    "PRICE_FOR_ALL_WORK",
    "TOTAL_PRICE",
]


@admin.register(Estimate)
class EstimateAdmin(admin.ModelAdmin):
    change_form_template = "admin/app_outlay/estimate_change.html"
    save_on_top = True
    list_per_page = 50
    list_display = (
        "name",
        "source_file",
        "currency",
    )
    search_fields = ("name",)
    readonly_fields = (
        "source_file",
        "source_sheet_index",
        "settings_data",
    )

    # Переопределим, чтобы убрать N+1 при списковом виде
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(
            _groups_count=Count("groups", distinct=False),
            _tc_links_count=Count("groups__techcard_links", distinct=True),
            _overhead_count=Count("overhead_cost_links", distinct=True),
        )

    # ---------- URLS: роутинг ----------
    # это нужно будет вынести отсюда. можно использовать DRF

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "<path:object_id>/api/export-excel/",
                self.admin_site.admin_view(self.api_export_excel),
                name="estimate_export_excel",
            ),
            # НОВЫЙ ЭНДПОИНТ - страница детального анализа
            path(
                "<path:object_id>/analysis/",
                self.admin_site.admin_view(self.analysis_view),
                name="estimate_analysis",
            ),
            # API для данных графиков
            path(
                "<path:object_id>/api/analysis-data/",
                self.admin_site.admin_view(self.api_analysis_data),
                name="estimate_analysis_data",
            ),
        ]
        return custom + urls

    def analysis_view(self, request, object_id: str):
        """
        Страница детального анализа сметы с графиками и декомпозицией цен.
        """
        from django.template.response import TemplateResponse

        est = self.get_object(request, object_id)
        if not est:
            messages.error(request, "Смета не найдена")
            return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))

        context = dict(
            self.admin_site.each_context(request),  # ← ИСПРАВЛЕНО
            title=f"Анализ сметы: {est.name}",
            estimate=est,
            has_data=est.groups.exists(),
        )

        return TemplateResponse(
            request, "admin/app_outlay/estimate_analysis.html", context
        )

    def api_analysis_data(self, request, object_id: str):
        """
        API: данные для графиков анализа сметы.

        ✓ База — по живым ценам (_base_costs_live).
        ✓ Продажи без НР — calc_for_tc(..., overhead_context=None).
        ✓ Продажи с НР — calc_for_tc(..., overhead_context=overhead_ctx из активных контейнеров сметы).
        ✓ Итог = продажи без НР + НР (фактически: сумма «с НР»).
        ✓ Группы — по ID (не склеиваются одноимённые).
        """
        from decimal import Decimal
        from django.http import JsonResponse
        from app_outlay.models import GroupTechnicalCardLink

        est = self.get_object(request, object_id)
        if not est:

            return JsonResponse({"ok": False, "error": "Смета не найдена"}, status=404)

        try:
            # --- связи Группа–Версия ТК
            tc_links = (
                GroupTechnicalCardLink.objects.filter(group__estimate=est)
                .select_related(
                    "technical_card_version", "technical_card_version__card", "group"
                )
                .order_by("group__order", "order")
            )
            tc_count = tc_links.count()

            if not tc_count:

                return JsonResponse(
                    {
                        "ok": True,
                        "has_data": False,
                        "message": "В смете нет привязанных техкарт",
                    }
                )

            # --------------------------------------------------------------------
            # 1) Сбор контекста НР из активных контейнеров (как в detail/export)
            # --------------------------------------------------------------------
            overhead_links = est.overhead_cost_links.filter(
                is_active=True
            ).select_related("overhead_cost_container")

            total_overhead_amt = Decimal("0")
            # суммы (amount * pct) для средневзвешенного разбиения МАТ/РАБ
            amount_weighted_mat_pct = Decimal("0")
            amount_weighted_work_pct = Decimal("0")

            overhead_breakdown = []
            for ol in overhead_links:
                amount = (
                    _dec(ol.snapshot_total_amount)
                    if ol.snapshot_total_amount is not None
                    else _dec(ol.overhead_cost_container.total_amount)
                )
                mat_pct = (
                    _dec(ol.snapshot_materials_percentage)
                    if ol.snapshot_materials_percentage is not None
                    else _dec(ol.overhead_cost_container.materials_percentage)
                )
                work_pct = (
                    _dec(ol.snapshot_works_percentage)
                    if ol.snapshot_works_percentage is not None
                    else _dec(ol.overhead_cost_container.works_percentage)
                )

                total_overhead_amt += amount
                amount_weighted_mat_pct += (mat_pct or 0) * amount
                amount_weighted_work_pct += (work_pct or 0) * amount

                overhead_breakdown.append(
                    {
                        "name": ol.overhead_cost_container.name,
                        "total": float(amount),
                        "materials_part": float(
                            amount * ((mat_pct or 0) / Decimal("100"))
                        ),
                        "works_part": float(
                            amount * ((work_pct or 0) / Decimal("100"))
                        ),
                        "materials_pct": float(mat_pct or 0),
                        "works_pct": float(work_pct or 0),
                    }
                )

            if total_overhead_amt > 0:
                avg_mat_pct = amount_weighted_mat_pct / total_overhead_amt  # 0..100
                avg_work_pct = amount_weighted_work_pct / total_overhead_amt
            else:
                avg_mat_pct = Decimal("0")
                avg_work_pct = Decimal("0")

            # Общая БАЗА (живые) — нужна для корректной работы движка с НР
            total_base_mat_live = Decimal("0")
            total_base_work_live = Decimal("0")
            for link in tc_links:
                ver = link.technical_card_version
                qty = _dec(getattr(link, "quantity", 1))
                base = _base_costs_live(ver)
                total_base_mat_live += _dec(base.mat) * qty
                total_base_work_live += _dec(base.work) * qty

            overhead_context = None
            if total_overhead_amt > 0:
                overhead_context = {
                    "total_base_mat": total_base_mat_live,
                    "total_base_work": total_base_work_live,
                    "overhead_amount": total_overhead_amt,
                    "overhead_mat_pct": avg_mat_pct,  # в процентах (0..100)
                    "overhead_work_pct": avg_work_pct,  # в процентах (0..100)
                }

            # --------------------------------------------------------------------
            # 2) Проходим по позициям и считаем (а) без НР и (б) с НР
            # --------------------------------------------------------------------
            total_base_mat = Decimal("0")
            total_base_work = Decimal("0")
            sales_mat_no_oh = Decimal("0")
            sales_work_no_oh = Decimal("0")
            sales_mat_with_oh = Decimal("0")
            sales_work_with_oh = Decimal("0")

            positions_data = []
            groups = {}  # агрегируем по group.id

            for link in tc_links:
                ver = link.technical_card_version
                if ver is None:
                    continue

                qty = _dec(getattr(link, "quantity", 1))

                # БАЗА — живые
                base = _base_costs_live(ver)
                line_base_mat = _dec(base.mat) * qty
                line_base_work = _dec(base.work) * qty
                total_base_mat += line_base_mat
                total_base_work += line_base_work

                # ПРОДАЖИ без НР
                calc0, _ = calc_for_tc(ver.card_id, float(qty), overhead_context=None)
                l0_mat = _dec(calc0.get("PRICE_FOR_ALL_MATERIAL", 0))
                l0_work = _dec(calc0.get("PRICE_FOR_ALL_WORK", 0))
                sales_mat_no_oh += l0_mat
                sales_work_no_oh += l0_work

                # ПРОДАЖИ с НР
                if overhead_context:
                    calc1, _ = calc_for_tc(
                        ver.card_id, float(qty), overhead_context=overhead_context
                    )
                    l1_mat = _dec(calc1.get("PRICE_FOR_ALL_MATERIAL", 0))
                    l1_work = _dec(calc1.get("PRICE_FOR_ALL_WORK", 0))
                else:
                    l1_mat, l1_work = l0_mat, l0_work
                sales_mat_with_oh += l1_mat
                sales_work_with_oh += l1_work

                positions_data.append(
                    {
                        "name": (ver.card.name or "")[:120],
                        "group": link.group.name,
                        "qty": float(qty),
                        "unit": ver.output_unit or "",
                        "base_mat": float(line_base_mat),
                        "base_work": float(line_base_work),
                        "final_mat": float(l0_mat),  # продажи без НР
                        "final_work": float(l0_work),  # продажи без НР
                        "total": float(l0_mat + l0_work),
                    }
                )

                # Группы по ID
                g = link.group
                if g.id not in groups:
                    groups[g.id] = {
                        "id": g.id,
                        "name": g.name,
                        "base_total": Decimal("0"),
                        "final_total": Decimal("0"),
                    }
                groups[g.id]["base_total"] += line_base_mat + line_base_work
                groups[g.id]["final_total"] += l0_mat + l0_work  # продажи без НР

            # --------------------------------------------------------------------
            # 3) Итоги/метрики
            # --------------------------------------------------------------------
            base_total = total_base_mat + total_base_work
            sales_total_no_oh = sales_mat_no_oh + sales_work_no_oh
            sales_total_with_oh = sales_mat_with_oh + sales_work_with_oh
            overhead_total_by_calc = (
                sales_total_with_oh - sales_total_no_oh
            )  # сколько реально добавлено НР

            avg_markup = (
                ((sales_total_no_oh - base_total) / base_total * 100)
                if base_total > 0
                else Decimal("0")
            )
            overhead_percent = (
                (overhead_total_by_calc / sales_total_no_oh * 100)
                if sales_total_no_oh > 0
                else Decimal("0")
            )

            # разбиение НР по Мат/Раб — из разницы «с НР» минус «без НР»
            oh_mat = sales_mat_with_oh - sales_mat_no_oh
            oh_work = sales_work_with_oh - sales_work_no_oh

            summary = {
                "base_materials": float(total_base_mat),
                "base_works": float(total_base_work),
                "base_total": float(base_total),
                "final_materials": float(sales_mat_no_oh),  # продажи без НР
                "final_works": float(sales_work_no_oh),  # продажи без НР
                "final_before_overhead": float(sales_total_no_oh),
                "overhead_total": float(overhead_total_by_calc),
                "final_with_overhead": float(sales_total_with_oh),
                "avg_markup_percent": float(avg_markup),
                "overhead_percent": float(overhead_percent),
                "positions_count": len(positions_data),
                "oh_split": {"materials": float(oh_mat), "works": float(oh_work)},
            }

            price_breakdown = {
                "labels": ["Себестоимость", "Продажи (без НР)", "Итог (с НР)"],
                "values": [
                    float(base_total),
                    float(sales_total_no_oh),
                    float(sales_total_with_oh),
                ],
            }

            groups_list = [
                {
                    "id": g["id"],
                    "name": g["name"],
                    "base_total": float(g["base_total"]),
                    "final_total": float(g["final_total"]),  # продажи без НР
                }
                for g in groups.values()
            ]
            groups_list.sort(key=lambda x: x["final_total"], reverse=True)

            materials_vs_works = {
                "labels": ["Материалы", "Работы"],
                "base": [float(total_base_mat), float(total_base_work)],
                "final": [float(sales_mat_no_oh), float(sales_work_no_oh)],  # без НР
            }
            materials_vs_works_after_oh = {
                "labels": ["Материалы", "Работы"],
                "values": [float(sales_mat_with_oh), float(sales_work_with_oh)],  # с НР
            }

            top_positions = sorted(
                positions_data, key=lambda x: x["total"], reverse=True
            )[:10]

            return JsonResponse(
                {
                    "ok": True,
                    "has_data": True,
                    "summary": summary,
                    "price_breakdown": price_breakdown,
                    "top_positions": top_positions,
                    "groups_distribution": groups_list,
                    "overhead_breakdown": overhead_breakdown,
                    "materials_vs_works": materials_vs_works,
                    "materials_vs_works_after_oh": materials_vs_works_after_oh,
                }
            )

        except Exception as e:
            import traceback

            return JsonResponse({"ok": False, "error": str(e)}, status=500)

    def api_export_excel(self, request, object_id: str):
        """
        Экспорт сметы с расчетами обратно в Excel.
        С детальным выводом через
        """
        # Получаем смету
        est = self.get_object(request, object_id)
        if not est:
            messages.error(request, "Смета не найдена")
            return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))
        # Проверяем наличие исходного файла
        if not est.source_file or not est.source_file.file:
            messages.error(request, "Исходный файл не найден")
            return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))
        # Проверяем наличие разметки
        if not hasattr(est.source_file, "markup"):
            messages.error(request, "Разметка файла не найдена")
            return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))

        try:
            markup = est.source_file.markup
            sheet_index = est.source_sheet_index or 0
            # Получаем схему колонок из разметки
            try:
                col_roles, _, _ = _SS().read_sheet_schema(markup, sheet_index)
            except Exception as e:
                schema = (
                    (markup.annotation or {})
                    .get("schema", {})
                    .get("sheets", {})
                    .get(str(sheet_index), {})
                )
                col_roles = schema.get("col_roles") or []
            if not col_roles:
                messages.error(request, "Не удалось определить структуру колонок")
                return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))

            # Загружаем Excel файл
            xlsx_path = est.source_file.file.path
            wb = load_workbook(xlsx_path)

            try:
                ws = wb.worksheets[sheet_index]
            except IndexError:
                ws = wb.active

            # ========== ПОДГОТОВКА КОНТЕКСТА НР + НДС ==========

            overhead_context = None
            overhead_links = est.overhead_cost_links.filter(
                is_active=True
            ).select_related("overhead_cost_container")

            if overhead_links.exists():
                total_overhead = Decimal("0")
                weighted_mat_pct = Decimal("0")
                weighted_work_pct = Decimal("0")

                for link in overhead_links:
                    amount = (
                        link.snapshot_total_amount
                        or link.overhead_cost_container.total_amount
                    )
                    mat_pct = (
                        link.snapshot_materials_percentage
                        or link.overhead_cost_container.materials_percentage
                    )
                    work_pct = (
                        link.snapshot_works_percentage
                        or link.overhead_cost_container.works_percentage
                    )

                    total_overhead += amount
                    weighted_mat_pct += mat_pct * amount
                    weighted_work_pct += work_pct * amount

                if total_overhead > 0:
                    avg_mat_pct = weighted_mat_pct / total_overhead
                    avg_work_pct = weighted_work_pct / total_overhead
                else:
                    avg_mat_pct = Decimal("0")
                    avg_work_pct = Decimal("0")

                # Рассчитываем общую базу из сохраненных ТК

                total_base_mat = Decimal("0")
                total_base_work = Decimal("0")

                tc_links = GroupTechnicalCardLink.objects.filter(
                    group__estimate=est
                ).select_related("technical_card_version")

                for idx, link in enumerate(tc_links[:5], 1):  # показываем первые 5
                    ver = link.technical_card_version
                    base = _base_costs_live(ver)
                    link_base_mat = base.mat * _dec(link.quantity)
                    link_base_work = base.work * _dec(link.quantity)
                    total_base_mat += link_base_mat
                    total_base_work += link_base_work

                if tc_links.count() > 5:
                    # считаем остальные
                    for link in tc_links[5:]:
                        ver = link.technical_card_version
                        base = _base_costs_live(ver)
                        total_base_mat += base.mat * _dec(link.quantity)
                        total_base_work += base.work * _dec(link.quantity)

                overhead_context = {
                    "total_base_mat": total_base_mat,
                    "total_base_work": total_base_work,
                    "overhead_amount": total_overhead,
                    "overhead_mat_pct": avg_mat_pct,
                    "overhead_work_pct": avg_work_pct,
                }

            # ========== ДОБАВЛЯЕМ НДС В КОНТЕКСТ ==========
            settings = est.settings_data or {}
            vat_active = settings.get("vat_active", False)
            vat_rate = settings.get("vat_rate", 20)

            # Если НР нет, но НДС есть — создаём контекст только для НДС
            if overhead_context is None and vat_active:
                overhead_context = {}

            # Добавляем НДС в контекст (если он есть)
            if overhead_context is not None:
                overhead_context["vat_active"] = vat_active
                overhead_context["vat_rate"] = vat_rate

            # Лог для отладки
            if vat_active:
                print(f"✅ Экспорт с НДС {vat_rate}%")
            else:
                print("ℹ Экспорт без НДС")

            # ========== ПОЛУЧАЕМ ВСЕ СОПОСТАВЛЕНИЯ ==========

            mappings = {}  # {row_index: {tc_version, quantity}}
            all_links = GroupTechnicalCardLink.objects.filter(
                group__estimate=est
            ).select_related("technical_card_version", "technical_card_version__card")
            links_with_row = 0
            links_without_row = 0

            for link in all_links:
                if link.source_row_index:
                    mappings[link.source_row_index] = {
                        "tc_version": link.technical_card_version,
                        "quantity": link.quantity,
                    }
                    links_with_row += 1
                else:
                    links_without_row += 1

            if not mappings:
                messages.warning(
                    request,
                    "⚠️ Нет сопоставлений для экспорта. Сначала выполните сопоставление и сохраните.",
                )
                wb.close()
                return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))

            # ========== ЗАПОЛНЯЕМ EXCEL ==========

            updated_count = 0
            # Создаем маппинг роль -> индекс колонки
            role_to_col = {}
            target_roles = [
                "QTY",
                "UNIT_PRICE_OF_MATERIAL",
                "UNIT_PRICE_OF_WORK",
                "UNIT_PRICE_OF_MATERIALS_AND_WORKS",
                "PRICE_FOR_ALL_MATERIAL",
                "PRICE_FOR_ALL_WORK",
                "TOTAL_PRICE",
            ]

            for idx, role in enumerate(col_roles):
                if role in target_roles:
                    role_to_col[role] = idx
                    col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + idx - 26)}"

            if not role_to_col:
                messages.error(request, "❌ Нет размеченных колонок для записи данных")
                wb.close()
                return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))

            # Проходим по всем сопоставлениям
            for idx, (row_index, mapping) in enumerate(mappings.items(), 1):
                tc_version = mapping["tc_version"]
                quantity = mapping["quantity"]

                if "QTY" in role_to_col:
                    col_idx = role_to_col["QTY"]
                    excel_row = row_index
                    excel_col = col_idx + 1

                    col_letter = (
                        chr(65 + col_idx)
                        if col_idx < 26
                        else f"A{chr(65 + col_idx - 26)}"
                    )
                    cell_address = f"{col_letter}{excel_row}"

                    try:
                        cell = ws.cell(row=excel_row, column=excel_col)
                        # Записываем количество как число
                        cell.value = float(quantity)
                        cell.number_format = "#,##0.000"
                    except Exception as e:
                        pass

                # Вызываем расчет с НР
                try:
                    calc, _ = calc_for_tc(
                        tc_version.card_id, quantity, overhead_context=overhead_context
                    )
                except Exception as e:
                    print(f"   ❌ ОШИБКА РАСЧЕТА: {e}")
                    continue

                # Записываем значения в соответствующие колонки
                cells_written = 0
                for role, value in calc.items():
                    if role in role_to_col:
                        col_idx = role_to_col[role]
                        excel_row = row_index
                        excel_col = col_idx + 1

                        col_letter = (
                            chr(65 + col_idx)
                            if col_idx < 26
                            else f"A{chr(65 + col_idx - 26)}"
                        )
                        cell_address = f"{col_letter}{excel_row}"

                        try:
                            cell = ws.cell(row=excel_row, column=excel_col)
                            # Записываем значение как число
                            cell.value = float(value)
                            cell.number_format = "#,##0.000"
                            cells_written += 1

                        except Exception as e:
                            print(f"      ❌ Ошибка записи в {cell_address}: {e}")

                if cells_written > 0:
                    updated_count += 1

                # Показываем только первые 5 детально
                if idx >= 5 and len(mappings) > 5:
                    # обрабатываем остальные без детального вывода
                    for row_index2, mapping2 in list(mappings.items())[5:]:
                        tc_version2 = mapping2["tc_version"]
                        quantity2 = mapping2["quantity"]
                        try:
                            calc2, _ = calc_for_tc(
                                tc_version2.card_id,
                                quantity2,
                                overhead_context=overhead_context,
                            )
                            for role, value in calc2.items():
                                if role in role_to_col:
                                    col_idx = role_to_col[role]
                                    excel_row = row_index2
                                    excel_col = col_idx + 1
                                    cell = ws.cell(row=excel_row, column=excel_col)
                                    cell.value = float(value)
                                    cell.number_format = "#,##0.000"
                            updated_count += 1
                        except Exception as e:
                            print(f"   ❌ Ошибка в строке {row_index2}: {e}")
                    break

            # ========== СОХРАНЯЕМ ФАЙЛ ==========

            temp_dir = tempfile.gettempdir()
            original_name = os.path.splitext(est.source_file.original_name)[0]
            output_filename = f"{original_name}_calculated.xlsx"
            temp_path = os.path.join(temp_dir, output_filename)

            wb.save(temp_path)
            wb.close()

            # Читаем и отправляем
            with open(temp_path, "rb") as f:
                file_content = f.read()

                response = HttpResponse(
                    file_content,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = (
                    f'attachment; filename="{output_filename}"'
                )

            # Удаляем временный файл
            try:
                os.unlink(temp_path)
            except Exception as e:
                print(f"   ⚠ Не удалось удалить временный файл: {e}")

            messages.success(
                request,
                f"✅ Экспорт выполнен успешно! Обновлено строк: {updated_count}",
            )
            return response

        except Exception as e:
            print("❌❌❌ КРИТИЧЕСКАЯ ОШИБКА ЭКСПОРТА ❌❌❌")
            print(f"Ошибка: {e}")
            messages.error(request, f"Ошибка экспорта: {e!r}")
            return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))

    # ---------- ВСПОМОГАТЕЛЬНОЕ: вытаскиваем индексы колонок по ролям ----------
    def _idxs(self, col_roles: list[str], role: str) -> list[int]:
        return [i for i, r in enumerate(col_roles or []) if r == role]

    def _cell(self, row: dict, idx: int) -> str:
        cells = row.get("cells") or []
        return (cells[idx] if 0 <= idx < len(cells) else "") or ""

    # ---------- ВСПОМОГАТЕЛЬНОЕ: детект ТК из ПОЛНЫХ строк ----------
    def _detect_tc_rows_from_rows(
        self,
        rows: list[dict],
        col_roles: list[str],
        unit_allow_set: set[str],
        require_qty: bool,
    ) -> list[dict]:
        name_cols = self._idxs(col_roles, "NAME_OF_WORK")
        unit_cols = self._idxs(col_roles, "UNIT")
        qty_cols = self._idxs(col_roles, "QTY")

        def first_text(row, idxs):
            for i in idxs:
                t = self._cell(row, i).strip()
                if t:
                    return t
            return ""

        def qty_ok(row):
            if not require_qty:
                return True
            for i in qty_cols:
                raw = self._cell(row, i).replace(" ", "").replace(",", ".")
                try:
                    if float(raw) > 0:
                        return True
                except Exception:
                    pass
            return False

        def normalize_unit(u: str) -> str:
            s = (u or "").lower().strip()
            s = s.replace("\u00b2", "2").replace("\u00b3", "3")
            compact = "".join(ch for ch in s if ch not in " .,")
            import re

            if re.fullmatch(r"(м\^?2|м2|квм|мкв|квадратн\w*метр\w*)", compact or ""):
                return "м2"
            if re.fullmatch(r"(м\^?3|м3|кубм|мкуб|кубическ\w*метр\w*)", compact or ""):
                return "м3"
            if re.fullmatch(r"(шт|штука|штуки|штук)", compact or ""):
                return "шт"
            if re.fullmatch(r"(пм|погм|погонныйметр|погонныхметров)", compact or ""):
                return "пм"
            if re.fullmatch(r"(компл|комплект|комплекта|комплектов)", compact or ""):
                return "компл"
            return compact

        tcs = []
        for row in rows:
            name = first_text(row, name_cols)
            unit_raw = first_text(row, unit_cols)
            unit = normalize_unit(unit_raw)
            if not name or not unit:
                continue
            if unit_allow_set and unit not in unit_allow_set:
                continue
            if not qty_ok(row):
                continue
            # qty для превью (может быть пустым)
            qty_val = first_text(row, qty_cols)
            tcs.append(
                {
                    "row_index": row.get("row_index"),
                    "name": name,
                    "unit": unit_raw,
                    "qty": qty_val,
                }
            )
        return tcs

    # ---------- ВСПОМОГАТЕЛЬНОЕ: собираем строки сопоставления (таблица) ----------

    def _collect_excel_candidates_from_rows(
        self, rows: list[dict], col_roles: list[str]
    ) -> list[dict]:
        name_cols = self._idxs(col_roles, "NAME_OF_WORK")
        unit_cols = self._idxs(col_roles, "UNIT")
        qty_cols = self._idxs(col_roles, "QTY")

        def first_text(row, idxs):
            for i in idxs:
                t = self._cell(row, i).strip()
                if t:
                    return t
            return ""

        # опциональные колонки: соберём индекс для каждой роли, если есть
        opt_idx = {
            rid: (self._idxs(col_roles, rid)[0] if self._idxs(col_roles, rid) else None)
            for rid in OPTIONAL_ROLE_IDS
        }

        out = []
        for row in rows:
            name = first_text(row, name_cols)
            unit = first_text(row, unit_cols)
            if not name and not unit:
                continue
            qty = first_text(row, qty_cols)
            excel_optional = {}
            for rid, ci in opt_idx.items():
                excel_optional[rid] = self._cell(row, ci) if ci is not None else ""
            out.append(
                {
                    "row_index": row.get("row_index"),
                    "name": name,
                    "unit": unit,
                    "qty": qty,
                    "excel_optional": excel_optional,
                }
            )
        return out

    # ---------- ВСПОМОГАТЕЛЬНОЕ: группы из annotation (толерантно) ----------
    def _load_groups_from_annotation(
        self, annotation: dict, sheet_i: int
    ) -> list[dict]:
        sheet_key = str(sheet_i)
        groups = []

        # Вариант A: annotation["schema"]["sheets"][sheet]["groups"]
        root = (annotation or {}).get("schema", {}).get("sheets", {}).get(sheet_key, {})
        if isinstance(root, dict) and isinstance(root.get("groups"), list):
            groups = root["groups"]

        # Фолбэк: annotation["groups"][sheet]
        if not groups:
            alt = (annotation or {}).get("groups", {}).get(sheet_key)
            if isinstance(alt, list):
                groups = alt
            elif isinstance(alt, dict) and isinstance(alt.get("items"), list):
                groups = alt["items"]

        # нормализация
        norm = []
        for g in groups or []:
            uid = g.get("uid") or g.get("id") or g.get("gid")
            if not uid:
                continue
            color = g.get("color") or "#e0f7fa"
            parent = g.get("parent_uid") or g.get("parent") or g.get("parentId")
            rows = g.get("rows") or g.get("ranges") or []
            rr = []
            for r in rows:
                if isinstance(r, (list, tuple)) and len(r) >= 2:
                    try:
                        rr.append([int(r[0]), int(r[1])])
                    except Exception:
                        pass
            norm.append(
                {
                    "uid": uid,
                    "name": g.get("name") or g.get("title") or "Группа",
                    "color": color,
                    "parent_uid": parent,
                    "rows": rr,
                }
            )
        return norm

    def _assign_tc_to_deepest_group(
        self, groups: list[dict], tcs: list[dict]
    ) -> tuple[list[dict], list[dict]]:
        """Возвращает (tree, loose). tree — список корневых групп с children и tcs."""
        by_id = {g["uid"]: g for g in groups}

        # глубина
        def depth(uid):
            d = 0
            cur = by_id.get(uid)
            while cur and cur.get("parent_uid"):
                d += 1
                cur = by_id.get(cur["parent_uid"])
            return d

        for g in groups:
            g["_depth"] = depth(g["uid"])

        def covered(g, row_idx: int) -> bool:
            for s, e in g.get("rows") or []:
                if s <= row_idx <= e:
                    return True
            return False

        # дерево групп
        children = {g["uid"]: [] for g in groups}
        roots = []
        for g in groups:
            pid = g.get("parent_uid")
            if pid and pid in children:
                children[pid].append(g)
            else:
                roots.append(g)

        # прикрепим ТК к самой глубокой накрывающей группе
        tcs_by_group = {g["uid"]: [] for g in groups}
        loose = []
        for tc in tcs:
            row_idx = tc.get("row_index")
            cands = [g for g in groups if covered(g, row_idx)]
            if cands:
                cands.sort(key=lambda x: x["_depth"])
                tcs_by_group[cands[-1]["uid"]].append(tc)
            else:
                loose.append(tc)

        # соберём дерево для вывода
        def build(u):
            node = by_id[u].copy()
            node["children"] = [build(ch["uid"]) for ch in children[u]]
            node["tcs"] = tcs_by_group[u]
            return node

        tree = [build(r["uid"]) for r in roots]
        return tree, loose

    # ---------- ПРЕДСТАВЛЕНИЯ: переопределенные представления ----------

    def render_change_form(
        self, request, context, add=False, change=False, form_url="", obj=None
    ):
        # на стандартный add оставляем дефолтное поведение
        if add or obj is None:
            return super().render_change_form(
                request, context, add, change, form_url, obj
            )

        # ========== ГРУППЫ И ТК ==========
        group_qs = Group.objects.filter(estimate=obj).order_by(
            "parent_id", "order", "id"
        )
        link_qs = (
            GroupTechnicalCardLink.objects.filter(group__estimate=obj)
            .select_related(
                "group", "technical_card_version", "technical_card_version__card"
            )
            .order_by("group_id", "order", "id")
        )

        # Создаём дефолтную группу если её нет
        if not group_qs.exists():
            Group.objects.get_or_create(
                estimate=obj,
                parent=None,
                defaults={"name": "Общий раздел", "order": 0},
            )
            group_qs = Group.objects.filter(estimate=obj).order_by(
                "parent_id", "order", "id"
            )

        # ========== НАКЛАДНЫЕ РАСХОДЫ ==========
        overhead_qs = (
            EstimateOverheadCostLink.objects.filter(estimate=obj)
            .select_related("overhead_cost_container")
            .order_by("order", "id")
        )

        # POST: валидируем и сохраняем формсеты
        if request.method == "POST":
            gfs = GroupFormSet(request.POST, queryset=group_qs, prefix="grp")
            lfs = LinkFormSet(request.POST, queryset=link_qs, prefix="lnk")

            # Формсет для накладных расходов
            from django.forms import modelformset_factory

            OverheadFormSet = modelformset_factory(
                EstimateOverheadCostLink,
                fields=["order", "overhead_cost_container", "is_active"],
                extra=1,
                can_delete=True,
            )
            ofs = OverheadFormSet(request.POST, queryset=overhead_qs, prefix="overhead")

            # Валидация всех формсетов
            if gfs.is_valid() and lfs.is_valid() and ofs.is_valid():
                with transaction.atomic():
                    gfs.save()
                    lfs.save()

                    # Сохраняем накладные расходы
                    for form in ofs.forms:
                        if form.cleaned_data and not form.cleaned_data.get(
                            "DELETE", False
                        ):
                            instance = form.save(commit=False)
                            instance.estimate = obj
                            instance.save()

                    # Удаляем помеченные
                    for form in ofs.deleted_forms:
                        if form.instance.pk:
                            form.instance.delete()

                self.message_user(
                    request, _("Изменения сохранены"), level=messages.SUCCESS
                )
                return HttpResponseRedirect(request.path)
            else:
                self.message_user(
                    request, _("Исправьте ошибки в форме"), level=messages.ERROR
                )
        else:
            gfs = GroupFormSet(queryset=group_qs, prefix="grp")
            lfs = LinkFormSet(queryset=link_qs, prefix="lnk")

            # Формсет для накладных расходов (GET)
            from django.forms import modelformset_factory

            OverheadFormSet = modelformset_factory(
                EstimateOverheadCostLink,
                fields=["order", "overhead_cost_container", "is_active"],
                extra=1,
                can_delete=True,
            )
            ofs = OverheadFormSet(queryset=overhead_qs, prefix="overhead")

        # ========== ПОСТРОЕНИЕ ДЕРЕВА ГРУПП ==========
        groups = list(group_qs)
        children = {}
        for g in groups:
            children.setdefault(g.parent_id, []).append(g)
        for lst in children.values():
            lst.sort(key=lambda x: (x.order, x.id))

        gform_by_id = {f.instance.pk: f for f in gfs.forms}
        lforms_by_gid = {}
        for f in lfs.forms:
            lforms_by_gid.setdefault(f.instance.group_id, []).append((f.instance, f))

        def build(parent_id):
            out = []
            for g in children.get(parent_id, []):
                out.append(
                    {
                        "group": g,
                        "group_form": gform_by_id.get(g.pk),
                        "links": lforms_by_gid.get(
                            g.pk, []
                        ),  # [(link_obj, link_form), ...]
                        "children": build(g.pk),
                    }
                )
            return out

        tree = build(None)

        # ========== КОНТЕКСТ ==========
        context.update(
            {
                "title": f"Смета: {obj.name}",
                "tree": tree,
                "group_formset": gfs,
                "link_formset": lfs,
                "overhead_formset": ofs,  # НОВОЕ
                "overhead_links": list(
                    overhead_qs
                ),  # НОВОЕ: для удобного доступа в шаблоне
                "list_url": reverse(
                    f"admin:{Estimate._meta.app_label}_{Estimate._meta.model_name}_changelist"
                ),
            }
        )
        return super().render_change_form(request, context, add, change, form_url, obj)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        extra = dict(extra_context or {})
        est = self.get_object(request, object_id)

        # Контекст по умолчанию
        table_sections: list[dict] = []
        optional_cols: list[dict] = []
        role_titles = ROLE_TITLES

        if (
            est
            and est.source_file_id
            and hasattr(est.source_file, "parse_result")
            and hasattr(est.source_file, "markup")
        ):
            pr = est.source_file.parse_result
            markup = est.source_file.markup
            sheet_i = est.source_sheet_index or 0

            # --- 1) Схема листа: роли колонок, allow-юниты, требование qty>0
            unit_allow_set = set()
            require_qty = False
            col_roles: list[str] = []

            def _normalize_unit(u: str) -> str:
                """Нормализация единиц измерения."""
                s = (u or "").lower().strip()
                s = s.replace("\u00b2", "2").replace("\u00b3", "3")
                compact = "".join(ch for ch in s if ch not in " .,")
                import re

                if re.fullmatch(
                    r"(м\^?2|м2|квм|мкв|квадратн\w*метр\w*)", compact or ""
                ):
                    return "м2"
                if re.fullmatch(
                    r"(м\^?3|м3|кубм|мкуб|кубическ\w*метр\w*)", compact or ""
                ):
                    return "м3"
                if re.fullmatch(r"(шт|штука|штуки|штук)", compact or ""):
                    return "шт"
                if re.fullmatch(
                    r"(пм|погм|погонныйметр|погонныхметров)", compact or ""
                ):
                    return "пм"
                if re.fullmatch(
                    r"(компл|комплект|комплекта|комплектов)", compact or ""
                ):
                    return "компл"
                return compact

            try:
                col_roles, unit_allow_set, require_qty = _SS().read_sheet_schema(
                    markup, sheet_i
                )
                unit_allow_set = set(
                    _normalize_unit(u) for u in (unit_allow_set or set())
                )
            except Exception:
                sch = (
                    (markup.annotation or {})
                    .get("schema", {})
                    .get("sheets", {})
                    .get(str(sheet_i), {})
                )
                col_roles = sch.get("col_roles") or []
                raw = (sch.get("unit_allow_raw") or "") if isinstance(sch, dict) else ""
                unit_allow_set = set()
                for part in (raw or "").split(","):
                    n = _normalize_unit(part)
                    if n:
                        unit_allow_set.add(n)
                require_qty = bool(sch.get("require_qty"))

            # --- 2) Загружаем ПОЛНЫЙ лист Excel
            xlsx_path = getattr(est.source_file.file, "path", None) or (
                (pr.data or {}).get("file") or {}
            ).get("path")
            if xlsx_path:
                rows_full = _load_full_sheet_rows_cached(xlsx_path, sheet_i)
            else:
                rows_full = ((pr.data or {}).get("sheets") or [{}])[sheet_i].get(
                    "rows"
                ) or []

            # --- 3) Детектируем кандидатов ТК
            tcs = self._detect_tc_rows_from_rows(
                rows_full, col_roles, unit_allow_set, require_qty
            )

            # --- 4) Группы/подгруппы из annotation и раскладка ТК
            groups = self._load_groups_from_annotation(markup.annotation or {}, sheet_i)
            tree, loose = self._assign_tc_to_deepest_group(groups, tcs)

            # --- 5) Собираем кандидатов для табличного вида
            allowed_rows = {tc["row_index"] for tc in tcs}
            excel_all = self._collect_excel_candidates_from_rows(rows_full, col_roles)
            # Фильтруем только те строки, которые прошли детект
            candidates_filtered = [
                it for it in excel_all if it["row_index"] in allowed_rows
            ]

            # --- 6) Опциональные колонки
            present_optional = [rid for rid in OPTIONAL_ROLE_IDS if rid in col_roles]
            optional_cols = [
                {"id": rid, "title": role_titles.get(rid, rid)}
                for rid in present_optional
            ]

            # Добавляем opt_values к каждому кандидату
            for it in candidates_filtered:
                raw = it.get("excel_optional") or {}
                it["opt_values"] = [raw.get(r["id"], "") for r in optional_cols]

            # --- 7) Собираем секции таблицы
            cand_by_row = {it["row_index"]: it for it in candidates_filtered}
            table_sections = []

            if groups:

                def _flatten(node: dict, parent_path: str | None = None):
                    path = node.get("name") or "Группа"
                    path = path if parent_path is None else f"{parent_path} / {path}"
                    items = []
                    for tc in node.get("tcs") or []:
                        ci = cand_by_row.get(tc["row_index"])
                        if ci:
                            items.append(ci)
                    if items:
                        table_sections.append(
                            {
                                "path": path,
                                "color": node.get("color") or "#eef",
                                "items": items,
                            }
                        )
                    for ch in node.get("children") or []:
                        _flatten(ch, path)

                for root in tree or []:
                    _flatten(root)

                # Остаток без группы
                loose_items = []
                for tc in loose or []:
                    ci = cand_by_row.get(tc["row_index"])
                    if ci:
                        loose_items.append(ci)
                if loose_items:
                    table_sections.append(
                        {"path": "Без группы", "color": "#f0f4f8", "items": loose_items}
                    )
            else:
                # Групп нет — одна секция со всеми найденными ТК
                if candidates_filtered:
                    table_sections = [
                        {
                            "path": "Без группы",
                            "color": "#f0f4f8",
                            "items": candidates_filtered,
                        }
                    ]

            # --- 8) Загружаем существующие сопоставления из БД
            existing_mappings = {}

            if est:
                links_qs = GroupTechnicalCardLink.objects.filter(
                    group__estimate=est
                ).select_related(
                    "group", "technical_card_version", "technical_card_version__card"
                )

                for link in links_qs:
                    if link.source_row_index:
                        existing_mappings[link.source_row_index] = {
                            "tc_id": link.technical_card_version.card_id,
                            "tc_name": link.technical_card_version.card.name,
                            "quantity": float(link.quantity),
                        }

        # URL для изменения ТК
        tc_change_url_zero = reverse(
            f"admin:{_TC._meta.app_label}_{_TC._meta.model_name}_change",
            args=[0],
        )

        # --- Финальный контекст (только нужное!)
        extra.update(
            {
                "table_sections": table_sections,
                "optional_cols": optional_cols,
                "calc_order_json": json.dumps(
                    [c["id"] for c in optional_cols], ensure_ascii=False
                ),
                "role_titles": role_titles,
                "table_colspan": 4 + len(optional_cols),
                "existing_mappings_json": json.dumps(
                    existing_mappings, ensure_ascii=False
                ),
                "tc_change_url_zero": tc_change_url_zero,
            }
        )

        return super().change_view(request, object_id, form_url, extra_context=extra)
