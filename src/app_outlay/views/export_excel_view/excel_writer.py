"""
Модуль записи данных в Excel файлы.

Следует принципам:
- Single Responsibility: только работа с Excel
- Clear API: явные методы для записи
- Error Handling: явная обработка ошибок записи
"""

import os
import tempfile
from typing import Dict, List

from openpyxl import load_workbook

from .exceptions import ExcelWriteError


class ExcelWriter:
    """
    Класс для записи рассчитанных данных в Excel.

    Ответственность:
    - Загрузка workbook
    - Создание маппинга роль → индекс колонки
    - Запись значений в ячейки
    - Форматирование ячеек
    - Сохранение файла
    """

    def __init__(self, xlsx_path: str, sheet_index: int):
        """
        Args:
            xlsx_path: Путь к исходному файлу Excel
            sheet_index: Индекс листа для записи
        """
        self.xlsx_path = xlsx_path
        self.sheet_index = sheet_index
        self.wb = None
        self.ws = None
        self.role_to_col: Dict[str, int] = {}

    def __enter__(self):
        """Context manager: загрузка workbook."""
        self.wb = load_workbook(self.xlsx_path)
        try:
            self.ws = self.wb.worksheets[self.sheet_index]
        except IndexError:
            self.ws = self.wb.active
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager: закрытие workbook."""
        if self.wb:
            self.wb.close()

    def setup_column_mapping(self, col_roles: List[str], target_roles: List[str]):
        """
        Создание маппинга роль → индекс колонки.

        Args:
            col_roles: Список ролей всех колонок из схемы
            target_roles: Список ролей для записи (QTY, UNIT_PRICE_OF_MATERIAL, ...)

        Example:
            >>> writer.setup_column_mapping(
            ...     ['NAME_OF_WORK', 'QTY', 'UNIT_PRICE_OF_MATERIAL'],
            ...     ['QTY', 'UNIT_PRICE_OF_MATERIAL']
            ... )
            >>> writer.role_to_col
            {'QTY': 1, 'UNIT_PRICE_OF_MATERIAL': 2}
        """
        self.role_to_col = {}
        for idx, role in enumerate(col_roles):
            if role in target_roles:
                self.role_to_col[role] = idx

    def write_value(
        self, row_index: int, role: str, value: float, number_format: str = "#,##0.000"
    ):
        """
        Запись значения в ячейку по роли.

        Args:
            row_index: Индекс строки (1-based, как в Excel)
            role: Роль колонки (например, 'QTY')
            value: Значение для записи
            number_format: Формат числа (default: "#,##0.000")

        Raises:
            ExcelWriteError: Если запись не удалась

        Example:
            >>> writer.write_value(10, 'QTY', 5.5)
            >>> writer.write_value(10, 'TOTAL_PRICE', 12345.67)
        """
        if role not in self.role_to_col:
            return  # Роль не размечена — пропускаем

        col_idx = self.role_to_col[role]
        excel_col = col_idx + 1

        try:
            cell = self.ws.cell(row=row_index, column=excel_col)
            cell.value = float(value)
            cell.number_format = number_format
        except Exception as e:
            col_letter = self._get_column_letter(col_idx)
            cell_address = f"{col_letter}{row_index}"
            raise ExcelWriteError(cell_address=cell_address, error=str(e))

    def write_calculated_row(self, row_index: int, calculations: Dict[str, float]):
        """
        Запись всех расчётных значений для строки.

        Args:
            row_index: Индекс строки
            calculations: Dict с результатами calc_for_tc
                {
                    'QTY': 5.0,
                    'UNIT_PRICE_OF_MATERIAL': 100.0,
                    'TOTAL_PRICE': 500.0,
                    ...
                }

        Returns:
            int: Количество записанных ячеек

        Example:
            >>> calc = {'QTY': 5.0, 'TOTAL_PRICE': 12345.67}
            >>> count = writer.write_calculated_row(10, calc)
            >>> count
            2
        """
        cells_written = 0

        for role, value in calculations.items():
            if role in self.role_to_col:
                self.write_value(row_index, role, value)
                cells_written += 1

        return cells_written

    def save_to_temp(self, original_filename: str) -> str:
        """
        Сохранение workbook во временный файл.

        Args:
            original_filename: Исходное имя файла (для генерации нового)

        Returns:
            str: Путь к временному файлу

        Example:
            >>> temp_path = writer.save_to_temp("estimate.xlsx")
            >>> temp_path
            '/tmp/estimate_calculated.xlsx'
        """
        temp_dir = tempfile.gettempdir()
        base_name = os.path.splitext(original_filename)[0]
        output_filename = f"{base_name}_calculated.xlsx"
        temp_path = os.path.join(temp_dir, output_filename)

        self.wb.save(temp_path)
        return temp_path

    @staticmethod
    def _get_column_letter(col_idx: int) -> str:
        """
        Получение буквы колонки из индекса.

        Args:
            col_idx: Индекс колонки (0-based)

        Returns:
            str: Буква колонки (A, B, ..., Z, AA, AB, ...)

        Example:
            >>> ExcelWriter._get_column_letter(0)
            'A'
            >>> ExcelWriter._get_column_letter(25)
            'Z'
            >>> ExcelWriter._get_column_letter(26)
            'AA'
        """
        if col_idx < 26:
            return chr(65 + col_idx)
        else:
            return f"A{chr(65 + col_idx - 26)}"
