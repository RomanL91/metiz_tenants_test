"""
Классы для чтения Excel файлов.

Следует принципам:
- Context Manager: безопасная работа с файлами
- Lazy Evaluation: ленивые свойства
- Single Responsibility: чёткое разделение ответственности
- Composition: ExcelSheetReader использует RowNormalizer
"""

import os
from functools import cached_property
from typing import Dict, Iterator, List, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .excel_cache import ExcelCacheManager
from .excel_normalizer import CellNormalizerStrategy, RowNormalizer
from .exceptions import (ExcelFileNotFoundError, ExcelReadError,
                         ExcelSheetNotFoundError)


class ExcelSheetReader:
    """
    Читатель отдельного листа Excel.

    Особенности:
    - Lazy loading данных
    - Автоматическое определение ширины
    - Кэширование результатов
    - Нормализация данных через Strategy

    Example:
        >>> reader = ExcelSheetReader(
        ...     path='/file.xlsx',
        ...     sheet_index=0,
        ...     use_cache=True
        ... )
        >>> rows = reader.read_all_rows()
        >>> len(rows)
        150
    """

    def __init__(
        self,
        path: str,
        sheet_index: int = 0,
        normalizer: RowNormalizer = None,
        use_cache: bool = True,
        cache_ttl: int = 600,
    ):
        """
        Args:
            path: Путь к файлу .xlsx
            sheet_index: Индекс листа (0-based)
            normalizer: Нормализатор строк (default: RowNormalizer())
            use_cache: Использовать кэширование
            cache_ttl: Время жизни кэша в секундах

        Raises:
            ExcelFileNotFoundError: Файл не найден
        """
        if not os.path.exists(path):
            raise ExcelFileNotFoundError(path=path)

        self.path = path
        self.sheet_index = sheet_index
        self.normalizer = normalizer or RowNormalizer()
        self.use_cache = use_cache
        self.cache_ttl = cache_ttl
        self._cache_manager = ExcelCacheManager()

    @cached_property
    def workbook(self) -> Workbook:
        """
        Lazy загрузка workbook.

        Workbook загружается только при первом обращении.
        """
        return load_workbook(self.path, data_only=True, read_only=True)

    @cached_property
    def worksheet(self) -> Worksheet:
        """
        Lazy получение worksheet.

        Raises:
            ExcelSheetNotFoundError: Лист не найден
        """
        try:
            return self.workbook.worksheets[self.sheet_index]
        except IndexError:
            total = len(self.workbook.worksheets)
            raise ExcelSheetNotFoundError(
                sheet_index=self.sheet_index, total_sheets=total
            )

    @cached_property
    def max_columns(self) -> int:
        """
        Автоматическое определение количества колонок.

        Анализирует первые 200 строк для определения ширины.

        Returns:
            int: Максимальное количество колонок
        """
        max_cols = 0
        sample_limit = min(200, self.worksheet.max_row)

        for row in self.worksheet.iter_rows(
            min_row=1, max_row=sample_limit, values_only=True
        ):
            # Проверяем что есть непустые ячейки
            if any(c not in (None, "", " ") for c in row):
                max_cols = max(max_cols, len(row))

        return max_cols if max_cols > 0 else (self.worksheet.max_column or 1)

    def read_all_rows(self) -> List[Dict[str, any]]:
        """
        Чтение всех строк листа с кэшированием.

        Returns:
            List[Dict]: Список строк в формате:
                [
                    {'row_index': 1, 'cells': ['val1', 'val2', ...]},
                    {'row_index': 2, 'cells': ['val1', 'val2', ...]},
                    ...
                ]

        Note:
            - row_index — 1-based (как в Excel)
            - cells — нормализованные значения

        Example:
            >>> rows = reader.read_all_rows()
            >>> rows[0]
            {'row_index': 1, 'cells': ['Header1', 'Header2', 'Header3']}
        """
        # Проверка кэша
        if self.use_cache:
            cached = self._cache_manager.get(self.path, self.sheet_index)
            if cached is not None:
                return cached

        # Чтение данных
        rows = self._read_rows_from_worksheet()

        # Сохранение в кэш
        if self.use_cache:
            self._cache_manager.set(self.path, self.sheet_index, rows, self.cache_ttl)

        return rows

    def _read_rows_from_worksheet(self) -> List[Dict[str, any]]:
        """
        Внутренний метод чтения строк из worksheet.

        Returns:
            List[Dict]: Нормализованные строки
        """
        rows = []
        max_cols = self.max_columns

        for idx, row in enumerate(
            self.worksheet.iter_rows(
                min_row=1, max_row=self.worksheet.max_row, values_only=True
            ),
            start=1,
        ):
            # Берём только нужное количество колонок
            cells = list(row)[:max_cols]

            # Нормализуем строку
            normalized_cells = self.normalizer.normalize_row(cells, max_cols)

            rows.append({"row_index": idx, "cells": normalized_cells})

        return rows

    def iter_rows(self, start_row: int = 1, end_row: int = None) -> Iterator[Dict]:
        """
        Итератор по строкам (для больших файлов).

        Args:
            start_row: Начальная строка (1-based)
            end_row: Конечная строка (None = до конца)

        Yields:
            Dict: Строка в формате {'row_index': int, 'cells': List[str]}

        Example:
            >>> for row in reader.iter_rows(start_row=10, end_row=20):
            ...     print(row['row_index'], row['cells'][0])
        """
        max_cols = self.max_columns
        end = end_row or self.worksheet.max_row

        for idx, row in enumerate(
            self.worksheet.iter_rows(min_row=start_row, max_row=end, values_only=True),
            start=start_row,
        ):
            cells = list(row)[:max_cols]
            normalized_cells = self.normalizer.normalize_row(cells, max_cols)

            yield {"row_index": idx, "cells": normalized_cells}

    def get_row(self, row_index: int) -> Optional[Dict]:
        """
        Получение конкретной строки.

        Args:
            row_index: Индекс строки (1-based)

        Returns:
            Dict или None: Строка или None если не найдена

        Example:
            >>> row = reader.get_row(10)
            >>> row['cells']
            ['cell1', 'cell2', 'cell3']
        """
        if row_index < 1 or row_index > self.worksheet.max_row:
            return None

        # Если данные закэшированы - берём оттуда
        if self.use_cache:
            all_rows = self.read_all_rows()
            for row in all_rows:
                if row["row_index"] == row_index:
                    return row
            return None

        # Иначе читаем напрямую
        for row in self.iter_rows(start_row=row_index, end_row=row_index):
            return row

        return None

    def invalidate_cache(self):
        """Инвалидация кэша для этого листа."""
        self._cache_manager.invalidate(self.path, self.sheet_index)

    def __enter__(self):
        """Context manager: вход."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager: выход с закрытием workbook."""
        if hasattr(self, "_workbook"):
            self.workbook.close()

    def __repr__(self):
        return f"<ExcelSheetReader(path={self.path}, sheet={self.sheet_index})>"


class ExcelWorkbookReader:
    """
    Читатель Excel workbook с управлением несколькими листами.

    Особенности:
    - Context Manager для безопасной работы
    - Lazy loading листов
    - Фабрика ExcelSheetReader'ов

    Example:
        >>> with ExcelWorkbookReader('/file.xlsx') as wb:
        ...     sheet0 = wb.get_sheet_reader(0)
        ...     rows = sheet0.read_all_rows()
    """

    def __init__(
        self,
        path: str,
        default_normalizer: RowNormalizer = None,
        use_cache: bool = True,
        cache_ttl: int = 600,
    ):
        """
        Args:
            path: Путь к файлу .xlsx
            default_normalizer: Нормализатор по умолчанию
            use_cache: Использовать кэширование
            cache_ttl: Время жизни кэша

        Raises:
            ExcelFileNotFoundError: Файл не найден
        """
        if not os.path.exists(path):
            raise ExcelFileNotFoundError(path=path)

        self.path = path
        self.default_normalizer = default_normalizer or RowNormalizer()
        self.use_cache = use_cache
        self.cache_ttl = cache_ttl
        self._sheet_readers: Dict[int, ExcelSheetReader] = {}

    def get_sheet_reader(
        self, sheet_index: int, normalizer: RowNormalizer = None
    ) -> ExcelSheetReader:
        """
        Получение reader'а для конкретного листа (Factory).

        Args:
            sheet_index: Индекс листа
            normalizer: Кастомный нормализатор (optional)

        Returns:
            ExcelSheetReader: Reader для листа

        Example:
            >>> reader = wb.get_sheet_reader(0)
            >>> rows = reader.read_all_rows()
        """
        # Кэшируем reader'ы
        if sheet_index not in self._sheet_readers:
            self._sheet_readers[sheet_index] = ExcelSheetReader(
                path=self.path,
                sheet_index=sheet_index,
                normalizer=normalizer or self.default_normalizer,
                use_cache=self.use_cache,
                cache_ttl=self.cache_ttl,
            )

        return self._sheet_readers[sheet_index]

    def read_sheet(self, sheet_index: int) -> List[Dict]:
        """
        Быстрое чтение листа (shortcut).

        Args:
            sheet_index: Индекс листа

        Returns:
            List[Dict]: Все строки листа

        Example:
            >>> rows = wb.read_sheet(0)
        """
        reader = self.get_sheet_reader(sheet_index)
        return reader.read_all_rows()

    def invalidate_all_cache(self):
        """Инвалидация кэша для всех листов."""
        cache_manager = ExcelCacheManager()
        cache_manager.invalidate(self.path)

    def __enter__(self):
        """Context manager: вход."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager: выход с очисткой reader'ов."""
        for reader in self._sheet_readers.values():
            if hasattr(reader, "workbook"):
                reader.workbook.close()
        self._sheet_readers.clear()

    def __repr__(self):
        return f"<ExcelWorkbookReader(path={self.path})>"
