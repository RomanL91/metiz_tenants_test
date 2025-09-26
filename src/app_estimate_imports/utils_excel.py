from __future__ import annotations
from typing import List, Dict, Any
from openpyxl import load_workbook


def load_sheet_rows_full(
    xlsx_path: str, sheet_index: int = 0, limit: int | None = None
) -> List[Dict[str, Any]]:
    """
    Возвращает ВСЕ строки листа Excel как [{"cells": [...], "row_index": int}, ...].
    read_only=True => память экономим; data_only=True => берём значения формул.
    limit — опционально ограничить сверху (для отладки).
    """
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[sheet_index]
    except IndexError:
        wb.close()
        return []

    out: List[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            else:
                # приводим к строке «по-человечески»
                cells.append(str(v))
        out.append({"cells": cells, "row_index": i})
        if limit and len(out) >= limit:
            break

    wb.close()
    return out
