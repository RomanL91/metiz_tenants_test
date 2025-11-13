from __future__ import annotations

import hashlib
from typing import BinaryIO, Optional

from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.styles import Color

from app_estimate_imports.models import ImportedEstimateFile, ParseResult


def compute_sha256(fobj: BinaryIO) -> str:
    h = hashlib.sha256()
    for chunk in iter(lambda: fobj.read(8192), b""):
        h.update(chunk)
    fobj.seek(0)
    return h.hexdigest()


def apply_tint(rgb_hex: str, tint: float) -> str:
    """
    Применяет tint (осветление/затемнение) к RGB цвету.

    Args:
        rgb_hex: Цвет в формате "RRGGBB" (без #)
        tint: Значение tint от -1.0 до 1.0
              > 0: осветление (смешивание с белым)
              < 0: затемнение (смешивание с чёрным)

    Returns:
        Цвет в формате "RRGGBB" после применения tint
    """
    if not tint or tint == 0.0:
        return rgb_hex

    # Парсим RGB компоненты
    r = int(rgb_hex[0:2], 16)
    g = int(rgb_hex[2:4], 16)
    b = int(rgb_hex[4:6], 16)

    # Применяем tint
    if tint > 0:
        # Осветление (смешивание с белым)
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    else:
        # Затемнение (смешивание с чёрным)
        r = int(r * (1 + tint))
        g = int(g * (1 + tint))
        b = int(b * (1 + tint))

    # Ограничиваем диапазон 0-255
    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))

    return f"{r:02X}{g:02X}{b:02X}"


def get_theme_color_rgb(workbook, theme_index: int) -> Optional[str]:
    """
    Извлекает RGB цвет из темы Excel по индексу.

    Args:
        workbook: Объект workbook openpyxl
        theme_index: Индекс цвета в теме (0-11)

    Returns:
        RGB в формате "RRGGBB" или None
    """
    try:
        # Стандартные цвета темы Office (если тема не найдена)
        # Индексы: 0=lt1, 1=dk1, 2=lt2, 3=dk2, 4=accent1, 5=accent2...
        default_theme_colors = {
            0: "FFFFFF",  # lt1 (Light 1)
            1: "000000",  # dk1 (Dark 1)
            2: "E7E6E6",  # lt2 (Light 2)
            3: "44546A",  # dk2 (Dark 2)
            4: "4472C4",  # accent1 (Blue)
            5: "ED7D31",  # accent2 (Orange)
            6: "A5A5A5",  # accent3 (Gray)
            7: "FFC000",  # accent4 (Yellow)
            8: "5B9BD5",  # accent5 (Light Blue)
            9: "70AD47",  # accent6 (Green)
        }

        # Пытаемся извлечь из темы workbook
        if hasattr(workbook, "_theme") and workbook._theme:
            theme = workbook._theme
            if hasattr(theme, "themeElements") and theme.themeElements:
                clr_scheme = theme.themeElements.clrScheme

                # Маппинг индекса на атрибут темы
                color_map = {
                    0: "lt1",
                    1: "dk1",
                    2: "lt2",
                    3: "dk2",
                    4: "accent1",
                    5: "accent2",
                    6: "accent3",
                    7: "accent4",
                    8: "accent5",
                    9: "accent6",
                }

                if theme_index in color_map:
                    color_name = color_map[theme_index]
                    if hasattr(clr_scheme, color_name):
                        theme_color_obj = getattr(clr_scheme, color_name)

                        # Извлекаем RGB из объекта цвета темы
                        if (
                            hasattr(theme_color_obj, "srgbClr")
                            and theme_color_obj.srgbClr
                        ):
                            rgb = theme_color_obj.srgbClr.val
                            if rgb and len(rgb) == 6:
                                return rgb.upper()
                        elif (
                            hasattr(theme_color_obj, "sysClr")
                            and theme_color_obj.sysClr
                        ):
                            # Системный цвет
                            rgb = theme_color_obj.sysClr.lastClr
                            if rgb and len(rgb) == 6:
                                return rgb.upper()

        # Fallback: используем стандартные цвета темы Office
        return default_theme_colors.get(theme_index)

    except (AttributeError, KeyError, IndexError):
        return None


def color_to_hex(color: Color, workbook) -> Optional[str]:
    """
    Конвертирует цвет openpyxl в hex формат с поддержкой всех типов цветов.

    Поддерживает:
    - RGB цвета (прямой #RRGGBB)
    - Theme цвета с tint (тематические из палитры Excel + осветление/затемнение)
    - Indexed цвета (индексированные из стандартной палитры)
    """
    if not color:
        return None

    # 1. Прямой RGB цвет
    if hasattr(color, "rgb") and color.rgb:
        rgb_str = str(color.rgb)
        # Если ARGB (8 символов), убираем альфа-канал
        if len(rgb_str) == 8:
            rgb_str = rgb_str[2:]

        if len(rgb_str) == 6 and all(c in "0123456789ABCDEFabcdef" for c in rgb_str):
            return f"#{rgb_str.upper()}"

    # 2. Theme цвет (с tint)
    if hasattr(color, "theme") and color.theme is not None:
        base_rgb = get_theme_color_rgb(workbook, color.theme)

        if base_rgb:
            # Применяем tint если есть
            tint_value = getattr(color, "tint", 0.0) or 0.0

            if tint_value != 0.0:
                final_rgb = apply_tint(base_rgb, tint_value)
            else:
                final_rgb = base_rgb

            return f"#{final_rgb}"

    # 3. Indexed цвет - пока не поддерживается
    # (можно добавить маппинг стандартной палитры позже)

    return None


def extract_cell_background_color(cell, workbook) -> Optional[str]:
    """
    Извлекает цвет фона ячейки с поддержкой разных типов заливки.

    Поддерживает:
    - PatternFill (сплошная заливка)
    - GradientFill (градиентная заливка - берём start_color)
    """
    if not cell.fill:
        return None

    fill = cell.fill

    # PatternFill - самый распространённый тип
    if hasattr(fill, "patternType") and fill.patternType:
        # Для сплошной заливки проверяем fgColor (foreground)
        if hasattr(fill, "fgColor") and fill.fgColor:
            return color_to_hex(fill.fgColor, workbook)
        # Также проверяем bgColor (background) как fallback
        if hasattr(fill, "bgColor") and fill.bgColor:
            return color_to_hex(fill.bgColor, workbook)

    # GradientFill - градиентная заливка
    if hasattr(fill, "start_color") and fill.start_color:
        return color_to_hex(fill.start_color, workbook)

    return None


def parse_excel_to_json(path: str) -> dict:
    """
    Парсит Excel файл с сохранением цветов ячеек.
    Читаем имена листов и первые 50 строк в виде текста + цвета фона.
    """
    # Загружаем С форматированием (НЕ read_only и НЕ data_only)
    wb = load_workbook(filename=path, data_only=False, read_only=False)

    result = {
        "file": {"path": path, "sheets": len(wb.worksheets)},
        "sheets": [],
        "extracted": {"estimate_name": ""},
    }

    for ws in wb.worksheets:
        rows = []
        for i, row in enumerate(ws.iter_rows()):

            cells = []
            colors = []

            for cell in row:
                # Значение ячейки
                value = str(cell.value) if cell.value is not None else ""
                cells.append(value)

                # Цвет фона ячейки (ИСПРАВЛЕНО с поддержкой theme+tint)
                bg_color = extract_cell_background_color(cell, wb)
                colors.append(bg_color)

            rows.append(
                {
                    "row_index": i + 1,
                    "cells": cells,
                    "colors": colors,
                }
            )

        result["sheets"].append({"name": ws.title, "rows": rows})

    return result


def save_imported_file(uploaded: UploadedFile) -> ImportedEstimateFile:
    obj = ImportedEstimateFile.objects.create(
        file=uploaded,
        original_name=uploaded.name,
        size_bytes=getattr(uploaded, "size", 0) or 0,
    )
    f = obj.file.open("rb")
    obj.sha256 = compute_sha256(f)
    obj.save(update_fields=["sha256"])
    return obj


@transaction.atomic
def parse_and_store(file_obj: ImportedEstimateFile) -> ParseResult:
    data = parse_excel_to_json(file_obj.file.path)
    estimate_name = (data.get("extracted", {}) or {}).get("estimate_name", "")[:255]
    pr, _ = ParseResult.objects.update_or_create(
        file=file_obj, defaults={"data": data, "estimate_name": estimate_name}
    )
    return pr
