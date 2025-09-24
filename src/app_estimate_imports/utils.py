import hashlib
from typing import BinaryIO

from openpyxl import load_workbook


def compute_sha256(fobj: BinaryIO) -> str:
    h = hashlib.sha256()
    for chunk in iter(lambda: fobj.read(8192), b""):
        h.update(chunk)
    fobj.seek(0)
    return h.hexdigest()


def parse_excel_to_json(path: str) -> dict:
    """
    МИНИМАЛЬНО: читаем имена листов и первые 50 строк в виде текста.
    Дальше усложним, когда определим эвристики/ML.
    """
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    result = {
        "file": {"path": path, "sheets": wb.worksheets.__len__()},
        "sheets": [],
        "extracted": {"estimate_name": ""},
    }
    for ws in wb.worksheets:
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 50:
                break
            cells = [str(v) if v is not None else "" for v in row]
            rows.append({"row_index": i + 1, "cells": cells})
        result["sheets"].append({"name": ws.title, "rows": rows})
    return result


def count_sheets_safely(path: str) -> int:
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        return len(wb.worksheets)
    except Exception:
        return 0
