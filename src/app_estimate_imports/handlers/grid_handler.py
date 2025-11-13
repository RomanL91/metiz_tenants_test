"""
Обработчик табличного представления (grid) для импортированного Excel.

Назначение
---------
- Отрисовать таблицу выбранного листа Excel в админке.
- Передать в шаблон служебные данные: роли колонок, заголовки, флаги показа, и т.п.
- Доверить бизнес-логику сервисам (markup/schema/group) и не смешивать её с UI.

Зависимости
-----------
- ROLE_DEFS: описание ролей колонок (id, заголовок, цвет, обязательность).
- load_sheet_rows_full: «ленивая» подгрузка полного листа из исходного XLSX при all=1.
- BaseHandler: базовая инфраструктура (admin, сервисы, сообщения, редиректы).

Важные замечания
----------------
- Метод show_grid ничего не вычисляет «по содержанию», он только собирает контекст.
- Поле unit_allow_raw в контексте формируется через ",".join(unit_allow_set).
  Порядок элементов в множестве не гарантирован — это ок для UI, но важно помнить.
"""

from __future__ import annotations

from typing import Any, Dict, List

from django.contrib import messages
from django.http import HttpRequest, HttpResponse
from django.template.response import TemplateResponse
from openpyxl import load_workbook

from app_estimate_imports.handlers.base_handler import BaseHandler
from app_estimate_imports.utils.constants import ROLE_DEFS


class GridHandler(BaseHandler):
    """
    Обработчик табличного представления.

    Содержит:
      - show_grid: основной entrypoint, рендерит grid.html с нужным контекстом.
      - _load_full_sheet_data: опциональная подгрузка всего листа (когда много строк).
      - _prepare_grid_context: сбор всех данных для шаблона.
      - _extract_column_headers: эвристика для заголовков колонок.
      - _get_role_definitions: передаёт в шаблон список доступных ролей.
    """

    def load_sheet_rows_full(
        self, xlsx_path: str, sheet_index: int = 0, limit: int | None = None
    ) -> List[Dict[str, Any]]:
        """
        Возвращает ВСЕ строки листа Excel как [{"cells": [...], "row_index": int}, ...].
        read_only=True => память экономим; data_only=True => берём значения формул.
        limit — опционально ограничить сверху (для отладки).
        """
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[sheet_index]
        except IndexError:
            wb.close()
            return []

        out: List[Dict[str, Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = []
            for v in row:
                if v is None:
                    cells.append("")
                else:
                    # приводим к строке «по-человечески»
                    cells.append(str(v))
            out.append({"cells": cells, "row_index": i})
            if limit and len(out) >= limit:
                break

        wb.close()
        return out

    def show_grid(self, request: HttpRequest, pk: int) -> HttpResponse:
        """
        Показывает табличное представление данных.

        Поток:
        1) Проверяем наличие ParseResult; если нет — сообщение и редирект.
        2) Определяем активный лист (?sheet=N), валидируем индекс.
        3) При необходимости подгружаем «полный» лист (?all=1).
        4) Собираем контекст и рендерим шаблон admin/app_estimate_imports/grid.html.

        :param    request: текущий HttpRequest.
        :param    pk: первичный ключ ImportedEstimateFile.

        :returns HttpResponse: HTML-страница с таблицей.
        """
        obj = self.get_object_or_error(request, pk)
        if not obj or not hasattr(obj, "parse_result"):
            messages.error(request, "Нет ParseResult")
            return self.redirect_back_or_change(request)

        pr = obj.parse_result
        sheets = pr.data.get("sheets") or []
        sheet_i = int(request.GET.get("sheet") or 0)

        if sheet_i < 0 or sheet_i >= len(sheets):
            sheet_i = 0

        sheet = sheets[sheet_i] if sheets else {"name": "Лист1", "rows": []}
        rows = sheet.get("rows") or []

        # Загрузка полных данных если запрошено
        show_all = request.GET.get("all") == "1"
        if show_all:
            rows = self._load_full_sheet_data(obj, sheet_i, rows)

        # Контекст для шаблона (включая роли, шапки, флаги и метаданные)
        context_data = self._prepare_grid_context(obj, request, sheet_i, sheets, rows)

        return TemplateResponse(
            request, "admin/app_estimate_imports/grid.html", context_data
        )

    def _load_full_sheet_data(self, obj, sheet_i: int, fallback_rows):
        """
        Загружает полные данные листа из файла.

        Когда в parse_result сохранена только «выдержка» (превью),
        этот метод достаёт все строки напрямую из XLSX (бывает много тысяч).

        :param obj: ImportedEstimateFile (ссылается на загруженный файл).
        :param sheet_i: индекс листа в книге Excel.
        :param fallback_rows: строки, которыми можно воспользоваться при ошибке.

        :returns list[dict]: список строк формата как в parse_result (rows).
                        Если файл недоступен/ошибка — вернёт fallback_rows.
        """
        try:

            xlsx_path = getattr(obj.file, "path", None) or (
                obj.parse_result.data.get("file") or {}
            ).get("path")

            if xlsx_path:
                return self.load_sheet_rows_full(xlsx_path, sheet_index=sheet_i)
        except Exception as e:
            self.add_error(f"Не удалось загрузить полный лист: {e!r}")

        return fallback_rows

    def _prepare_grid_context(
        self, obj, request: HttpRequest, sheet_i: int, sheets, rows
    ):
        """Подготавливает контекст для шаблона"""
        max_cols = max((len(r.get("cells") or []) for r in rows), default=0)
        cols = list(range(max_cols))
        sheet_names = [s.get("name") or f"Лист {i+1}" for i, s in enumerate(sheets)]

        # Получение схемы
        markup = self.markup_service.ensure_markup_exists(obj)
        col_roles, unit_allow_set, require_qty = self.schema_service.get_schema_config(
            markup, sheet_i
        )

        # Дополнение ролей до нужной длины
        if len(col_roles) < max_cols:
            col_roles = (col_roles + ["NONE"] * (max_cols - len(col_roles)))[:max_cols]

        # Заголовки колонок
        col_headers = self._extract_column_headers(rows, max_cols)

        # Проверяем наличие markup с данными (не пустой)
        has_valid_markup = hasattr(obj, "markup") and obj.markup.annotation

        return dict(
            self.admin.admin_site.each_context(request),
            title=f"Таблица: {obj.original_name}",
            file=obj,
            sheet_index=sheet_i,
            sheet_names=sheet_names,
            rows=rows,
            cols=cols,
            role_defs=self._get_role_definitions(),
            col_roles=col_roles,
            col_headers=col_headers,
            show_all=request.GET.get("all") == "1",
            total_rows=len(rows),
            unit_allow_raw=",".join(unit_allow_set),
            require_qty=require_qty,
            has_markup=has_valid_markup,  # Добавлено для кнопки "Создать смету"
        )

    def _extract_column_headers(self, rows, max_cols: int):
        """Извлекает заголовки колонок из первых строк"""
        col_headers = []
        for col_idx in range(max_cols):
            header = ""
            for row in rows[:8]:  # Ищем в первых 8 строках
                cells = row.get("cells") or []
                if col_idx < len(cells):
                    val = (cells[col_idx] or "").strip()
                    if val:
                        header = val
                        break
            col_headers.append(header)
        return col_headers

    def _get_role_definitions(self):
        """Возвращает определения ролей колонок"""

        return ROLE_DEFS
